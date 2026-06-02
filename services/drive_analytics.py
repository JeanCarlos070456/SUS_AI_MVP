from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
import warnings
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
DEFAULT_CACHE_DIR = SCRIPTS_DIR / "drive_cache"

EXCEL_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/vnd.ms-excel.sheet.macroenabled.12",
}

PDF_MIME = "application/pdf"
CSV_MIME = "text/csv"

SUPPORTED_ANALYSIS_MIMES = {
    PDF_MIME,
    CSV_MIME,
    *EXCEL_MIMES,
}

ANALYTIC_KEYWORDS = {
    "imd",
    "indicador",
    "indicadores",
    "meta",
    "metas",
    "resultado",
    "resultados",
    "desempenho",
    "monitoramento",
    "avaliacao",
    "avaliação",
    "custo",
    "custos",
    "producao",
    "produção",
    "apuracao",
    "apuração",
    "apura",
    "apurasus",
    "total",
    "media",
    "média",
    "percentual",
    "taxa",
    "valor",
    "variação",
    "variacao",
    "unidade",
    "regiao",
    "região",
    "srs",
    "hospital",
    "policlinica",
    "policlínica",
    "aps",
    "dashboard",
    "relatorio",
    "relatório",
    "analitico",
    "analítico",
    "status",
    "critico",
    "crítico",
    "alerta",
    "baixo",
    "alto",
    "pior",
    "melhor",
}

NOISE_VALUES = {
    "",
    "none",
    "nan",
    "nat",
    "unnamed",
    "unnamed:",
    "total geral",
}

REGION_RULES = {
    "norte": {
        "strong_positive": [
            "rgqc regiao norte",
            "rgqc - regiao norte",
            "rgqc região norte",
            "rgqc - região norte",
            "srs norte",
            "srsno",
            "regiao norte",
            "região norte",
        ],
        "weak_positive": [
            " norte ",
        ],
        "negative": [
            "asa norte",
            "policlinica asa norte",
            "policlinica asa norte",
            "regiao central",
            "região central",
            "rgqc regiao central",
            "rgqc - regiao central",
            "srs central",
            "srsce",
            "central - asa norte",
        ],
    },
    "oeste": {
        "strong_positive": [
            "rgqc regiao oeste",
            "rgqc - regiao oeste",
            "rgqc região oeste",
            "rgqc - região oeste",
            "srs oeste",
            "srsoe",
            "regiao oeste",
            "região oeste",
        ],
        "weak_positive": [
            " oeste ",
        ],
        "negative": [
            "sudoeste",
            "regiao sudoeste",
            "região sudoeste",
            "rgqc regiao sudoeste",
            "rgqc - regiao sudoeste",
            "srs sudoeste",
        ],
    },
    "sul": {
        "strong_positive": [
            "rgqc regiao sul",
            "rgqc - regiao sul",
            "rgqc região sul",
            "rgqc - região sul",
            "srs sul",
            "srssul",
            "regiao sul",
            "região sul",
        ],
        "weak_positive": [
            " sul ",
        ],
        "negative": [
            "centro-sul",
            "centro sul",
            "regiao centro sul",
            "região centro sul",
            "regiao centro-sul",
            "região centro-sul",
            "srs centro-sul",
            "srscs",
            "lago sul",
        ],
    },
    "centro-sul": {
        "strong_positive": [
            "rgqc regiao centro sul",
            "rgqc - regiao centro sul",
            "rgqc regiao centro-sul",
            "rgqc - regiao centro-sul",
            "rgqc região centro sul",
            "rgqc - região centro sul",
            "rgqc região centro-sul",
            "rgqc - região centro-sul",
            "srs centro-sul",
            "srs centro sul",
            "srscs",
            "regiao centro-sul",
            "região centro-sul",
            "regiao centro sul",
            "região centro sul",
        ],
        "weak_positive": [
            " centro sul ",
            " centro-sul ",
        ],
        "negative": [
            "regiao sul",
            "região sul",
            "srs sul",
            "lago sul",
        ],
    },
    "central": {
        "strong_positive": [
            "rgqc regiao central",
            "rgqc - regiao central",
            "rgqc região central",
            "rgqc - região central",
            "srs central",
            "srsce",
            "regiao central",
            "região central",
        ],
        "weak_positive": [
            " central ",
        ],
        "negative": [
            "centro-sul",
            "centro sul",
            "regiao centro-sul",
            "região centro-sul",
        ],
    },
    "leste": {
        "strong_positive": [
            "rgqc regiao leste",
            "rgqc - regiao leste",
            "rgqc região leste",
            "rgqc - região leste",
            "srs leste",
            "srsle",
            "regiao leste",
            "região leste",
        ],
        "weak_positive": [
            " leste ",
        ],
        "negative": [],
    },
    "sudoeste": {
        "strong_positive": [
            "rgqc regiao sudoeste",
            "rgqc - regiao sudoeste",
            "rgqc região sudoeste",
            "rgqc - região sudoeste",
            "srs sudoeste",
            "srssudoeste",
            "regiao sudoeste",
            "região sudoeste",
            "sudoeste",
        ],
        "weak_positive": [],
        "negative": [
            "regiao oeste",
            "região oeste",
            "srs oeste",
            "srsoe",
        ],
    },
}


@dataclass
class FileInspection:
    ok: bool
    file_name: str
    file_id: str
    path: str
    mime_type: str
    local_path: str = ""
    kind: str = ""
    message: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)
    sheets: list[dict[str, Any]] = field(default_factory=list)
    text_preview: str = ""
    analytic_summary: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


@dataclass
class DriveAnalysisResult:
    ok: bool
    requires_drive: bool
    catalog_available: bool
    found_sources: bool
    message: str
    question: str
    selected_sources: list[dict[str, Any]] = field(default_factory=list)
    inspections: list[dict[str, Any]] = field(default_factory=list)
    text: str = ""
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "requires_drive": self.requires_drive,
            "catalog_available": self.catalog_available,
            "found_sources": self.found_sources,
            "message": self.message,
            "question": self.question,
            "selected_sources": self.selected_sources,
            "inspections": self.inspections,
            "text": self.text,
            "errors": self.errors,
        }


def _bootstrap_import_path() -> None:
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[/\\_]+", " ", text)
    text = re.sub(r"[-–—]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _tokenize(value: Any) -> list[str]:
    text = _normalize_text(value)
    return re.findall(r"[a-zA-Z0-9]+", text)


def _safe_get(item: dict[str, Any], key: str, default: str = "") -> str:
    value = item.get(key, default)
    return "" if value is None else str(value)


def _sanitize_filename(name: str) -> str:
    name = name.strip() or "arquivo_drive"
    name = re.sub(r"[^\w\-.() áàâãéêíóôõúçÁÀÂÃÉÊÍÓÔÕÚÇ]+", "_", name, flags=re.UNICODE)
    name = re.sub(r"_+", "_", name)
    return name[:180]


def _ensure_cache_dir(cache_dir: str | Path | None = None) -> Path:
    path = Path(cache_dir) if cache_dir else DEFAULT_CACHE_DIR
    path.mkdir(parents=True, exist_ok=True)
    return path


def _write_bytes_to_cache(
    *,
    file_bytes: bytes,
    file_name: str,
    file_id: str,
    cache_dir: str | Path | None = None,
) -> Path:
    cache_path = _ensure_cache_dir(cache_dir)
    safe_name = _sanitize_filename(file_name)

    if "." not in Path(safe_name).name:
        safe_name = f"{safe_name}.bin"

    local_path = cache_path / f"{file_id}_{safe_name}"
    local_path.write_bytes(file_bytes)
    return local_path


def _extract_years(message: str) -> list[str]:
    return re.findall(r"\b(20\d{2})\b", message or "")


def _contains_phrase(blob: str, phrase: str) -> bool:
    blob_norm = f" {_normalize_text(blob)} "
    phrase_norm = f" {_normalize_text(phrase)} "
    return phrase_norm in blob_norm


def _infer_region_hint(message: str) -> str:
    """
    Detecta região/unidade principal na pergunta.

    Agora a inferência é mais rígida:
    - "região norte" = Região Norte, não Asa Norte.
    - "região oeste" = Região Oeste, não Sudoeste.
    - "região sul" = Região Sul, não Centro-Sul/Lago Sul.
    """
    text = f" {_normalize_text(message)} "

    explicit_patterns = [
        ("centro-sul", ["regiao centro sul", "regiao centro-sul", "srs centro sul", "srs centro-sul", "centro sul", "centro-sul"]),
        ("sudoeste", ["regiao sudoeste", "srs sudoeste", "sudoeste"]),
        ("central", ["regiao central", "srs central", "regional central"]),
        ("leste", ["regiao leste", "srs leste", "regional leste"]),
        ("norte", ["regiao norte", "srs norte", "regional norte"]),
        ("oeste", ["regiao oeste", "srs oeste", "regional oeste"]),
        ("sul", ["regiao sul", "srs sul", "regional sul"]),
    ]

    for canonical, patterns in explicit_patterns:
        for pattern in patterns:
            if _contains_phrase(text, pattern):
                return canonical

    # Fallback para perguntas curtas, ex.: "RGQC norte 2025".
    fallback_patterns = [
        ("centro-sul", ["centro sul", "centro-sul"]),
        ("sudoeste", ["sudoeste"]),
        ("central", ["central"]),
        ("leste", ["leste"]),
        ("norte", ["norte"]),
        ("oeste", ["oeste"]),
        ("sul", ["sul"]),
    ]

    for canonical, patterns in fallback_patterns:
        for pattern in patterns:
            if _contains_phrase(text, pattern):
                return canonical

    units = [
        "hmib",
        "hab",
        "hcb",
        "hub",
        "hsvp",
        "igesdf",
        "samu",
        "hrc",
        "hrbz",
        "hran",
        "hrgu",
        "hrl",
        "hrpl",
    ]

    for unit in units:
        if _contains_phrase(text, unit):
            return unit

    return ""


def _source_blob(source: dict[str, Any]) -> str:
    fields = [
        source.get("name", ""),
        source.get("path", ""),
        source.get("category_detected", ""),
        source.get("area_detected", ""),
        source.get("year_detected", ""),
        source.get("region_or_unit_detected", ""),
        source.get("mimeType", ""),
        source.get("extension", ""),
    ]
    return _normalize_text(" ".join(str(f) for f in fields))


def _question_has_explicit_region(question: str) -> bool:
    text = _normalize_text(question)
    return any(
        phrase in text
        for phrase in [
            "regiao norte",
            "regiao oeste",
            "regiao sul",
            "regiao central",
            "regiao leste",
            "regiao sudoeste",
            "regiao centro sul",
            "regiao centro-sul",
            "srs norte",
            "srs oeste",
            "srs sul",
            "srs central",
            "srs leste",
            "srs sudoeste",
            "srs centro sul",
            "srs centro-sul",
        ]
    )


def _territorial_score_adjustment(source: dict[str, Any], question: str) -> tuple[int, list[str]]:
    region_hint = _infer_region_hint(question)

    if not region_hint or region_hint not in REGION_RULES:
        return 0, []

    rules = REGION_RULES[region_hint]
    blob = f" {_source_blob(source)} "
    explanation: list[str] = []
    adjustment = 0
    has_strong_match = False

    for pattern in rules["strong_positive"]:
        if _contains_phrase(blob, pattern):
            adjustment += 130
            has_strong_match = True
            explanation.append(f"+território forte:{pattern}")
            break

    if not has_strong_match:
        for pattern in rules["weak_positive"]:
            if _contains_phrase(blob, pattern):
                adjustment += 25
                explanation.append(f"+território fraco:{pattern}")
                break

    for pattern in rules["negative"]:
        if _contains_phrase(blob, pattern):
            adjustment -= 140
            explanation.append(f"-território incompatível:{pattern}")
            break

    # Regra dura: se a pergunta pede explicitamente uma Região de Saúde e o arquivo está em outra pasta RGQC,
    # penaliza forte. Isso resolve "Região Norte" puxando "Asa Norte" da Região Central.
    if _question_has_explicit_region(question):
        rgqc_region_patterns = [
            "rgqc regiao norte",
            "rgqc regiao oeste",
            "rgqc regiao sul",
            "rgqc regiao central",
            "rgqc regiao leste",
            "rgqc regiao sudoeste",
            "rgqc regiao centro sul",
            "rgqc regiao centro-sul",
        ]

        matching_rgqc_region = [p for p in rgqc_region_patterns if _contains_phrase(blob, p)]

        if matching_rgqc_region:
            expected = {
                "norte": ["rgqc regiao norte"],
                "oeste": ["rgqc regiao oeste"],
                "sul": ["rgqc regiao sul"],
                "central": ["rgqc regiao central"],
                "leste": ["rgqc regiao leste"],
                "sudoeste": ["rgqc regiao sudoeste"],
                "centro-sul": ["rgqc regiao centro sul", "rgqc regiao centro-sul"],
            }.get(region_hint, [])

            if not any(_contains_phrase(blob, p) for p in expected):
                adjustment -= 220
                explanation.append("-pasta_rgqc_de_outra_região")

    return adjustment, explanation


def _adjust_source_score(source: dict[str, Any], question: str) -> int:
    base_score = int(source.get("score") or 0)
    score = base_score

    blob = _source_blob(source)
    name = _normalize_text(source.get("name", ""))
    mime = _safe_get(source, "mimeType")
    years = _extract_years(question)

    if mime in SUPPORTED_ANALYSIS_MIMES:
        score += 10
    else:
        score -= 20

    for year in years:
        if str(source.get("year_detected", "")) == year:
            score += 25
        elif year in blob:
            score += 10

    territorial_adjustment, _ = _territorial_score_adjustment(source, question)
    score += territorial_adjustment

    category = _normalize_text(source.get("category_detected", ""))
    if category in {"imd", "rgqc", "apura_sus", "excel", "documento_pdf"}:
        score += 8

    if "thumbs db" in name:
        score -= 200

    return score


def _rerank_sources_for_question(
    sources: list[dict[str, Any]],
    question: str,
    *,
    limit: int = 3,
) -> list[dict[str, Any]]:
    adjusted: list[dict[str, Any]] = []

    for source in sources:
        item = dict(source)
        item["adjusted_score"] = _adjust_source_score(item, question)
        territorial_adjustment, territorial_notes = _territorial_score_adjustment(item, question)
        item["territorial_adjustment"] = territorial_adjustment
        item["territorial_notes"] = territorial_notes
        adjusted.append(item)

    adjusted.sort(
        key=lambda x: (
            int(x.get("adjusted_score") or 0),
            str(x.get("modifiedTime", "")),
            int(x.get("size_bytes") or 0),
        ),
        reverse=True,
    )

    return adjusted[: max(limit, 0)]


def _inspection_to_dict(inspection: FileInspection) -> dict[str, Any]:
    return {
        "ok": inspection.ok,
        "file_name": inspection.file_name,
        "file_id": inspection.file_id,
        "path": inspection.path,
        "mime_type": inspection.mime_type,
        "local_path": inspection.local_path,
        "kind": inspection.kind,
        "message": inspection.message,
        "metadata": inspection.metadata,
        "sheets": inspection.sheets,
        "text_preview": inspection.text_preview,
        "analytic_summary": inspection.analytic_summary,
        "errors": inspection.errors,
    }


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return str(value).strip()


def _is_noise_text(value: str) -> bool:
    norm = _normalize_text(value)
    if norm in NOISE_VALUES:
        return True
    if norm.startswith("unnamed"):
        return True
    return False


def _is_numeric_like(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True

    text = str(value).strip()
    if not text:
        return False

    text = text.replace("R$", "").replace("%", "").replace(".", "").replace(",", ".").strip()

    try:
        float(text)
        return True
    except Exception:
        return False


def _detect_numeric_type(value: Any) -> str:
    text = str(value).strip()

    if "%" in text:
        return "percentual"
    if "R$" in text or "r$" in text.lower():
        return "monetario"
    if _is_numeric_like(value):
        return "numerico"
    return ""


def _row_to_clean_values(row: list[Any]) -> list[str]:
    values: list[str] = []
    for value in row:
        text = _cell_to_text(value)
        if not _is_noise_text(text):
            values.append(text)
    return values


def _score_text_for_question(text: str, question: str) -> tuple[int, list[str]]:
    norm_text = _normalize_text(text)
    question_terms = set(_tokenize(question))
    keyword_terms = {_normalize_text(k) for k in ANALYTIC_KEYWORDS}

    score = 0
    hits: list[str] = []

    for term in question_terms:
        if len(term) <= 2:
            continue
        if term in norm_text:
            score += 5
            hits.append(term)

    for kw in keyword_terms:
        if len(kw) <= 2:
            continue
        if kw in norm_text:
            score += 3
            hits.append(kw)

    years = _extract_years(question)
    for year in years:
        if year in norm_text:
            score += 8
            hits.append(year)

    region_hint = _infer_region_hint(question)
    if region_hint and _contains_phrase(norm_text, region_hint):
        score += 12
        hits.append(region_hint)

    return score, sorted(set(hits))


def _rank_sheet_name(sheet_name: str, question: str) -> tuple[int, list[str]]:
    score, hits = _score_text_for_question(sheet_name, question)

    name = _normalize_text(sheet_name)

    priority_patterns = {
        "dashboard": 12,
        "imd": 12,
        "estratificado": 10,
        "relatorio analitico": 10,
        "relatorio": 6,
        "analitico": 6,
        "aps": 7,
        "custo": 7,
        "apuracao": 7,
        "producao": 6,
        "resultado": 5,
    }

    for pattern, weight in priority_patterns.items():
        if pattern in name:
            score += weight
            hits.append(pattern)

    return score, sorted(set(hits))


def _scan_excel_sheet_cells(
    *,
    ws,
    sheet_name: str,
    question: str,
    max_scan_rows: int,
    max_scan_cols: int,
    top_rows_limit: int = 12,
) -> dict[str, Any]:
    non_empty_cells = 0
    numeric_cells = 0
    percent_cells = 0
    monetary_cells = 0
    keyword_hits_counter: Counter[str] = Counter()
    candidate_rows: list[dict[str, Any]] = []
    dense_rows = 0

    max_row = min(ws.max_row or 0, max_scan_rows)
    max_col = min(ws.max_column or 0, max_scan_cols)

    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        start=1,
    ):
        row_values = list(row)
        clean_values = _row_to_clean_values(row_values)

        if len(clean_values) >= 3:
            dense_rows += 1

        row_text = " | ".join(clean_values)
        row_score, hits = _score_text_for_question(row_text, question)

        row_numeric_count = 0
        row_numeric_examples: list[str] = []

        for value in row_values:
            text = _cell_to_text(value)
            if not text:
                continue

            non_empty_cells += 1

            numeric_type = _detect_numeric_type(value)
            if numeric_type:
                numeric_cells += 1
                row_numeric_count += 1

                if len(row_numeric_examples) < 6:
                    row_numeric_examples.append(text)

                if numeric_type == "percentual":
                    percent_cells += 1
                elif numeric_type == "monetario":
                    monetary_cells += 1

        for hit in hits:
            keyword_hits_counter[hit] += 1

        if row_score > 0 or row_numeric_count >= 3:
            candidate_rows.append(
                {
                    "row_index": row_idx,
                    "score": row_score + min(row_numeric_count, 5),
                    "hits": hits,
                    "numeric_count": row_numeric_count,
                    "numeric_examples": row_numeric_examples,
                    "values": clean_values[:18],
                    "text": row_text[:1000],
                }
            )

    candidate_rows.sort(
        key=lambda x: (int(x.get("score") or 0), int(x.get("numeric_count") or 0)),
        reverse=True,
    )

    structure_type = "tabela_ou_bloco"
    if max_row > 0 and non_empty_cells > 0:
        avg_non_empty_per_row = non_empty_cells / max_row
        if avg_non_empty_per_row < 4 and dense_rows < max(3, max_row * 0.15):
            structure_type = "dashboard_esparso"
        elif dense_rows >= max(5, max_row * 0.35):
            structure_type = "tabela_densa"
    else:
        avg_non_empty_per_row = 0

    return {
        "sheet_name": sheet_name,
        "scan_rows": max_row,
        "scan_cols": max_col,
        "non_empty_cells": non_empty_cells,
        "numeric_cells": numeric_cells,
        "percent_cells": percent_cells,
        "monetary_cells": monetary_cells,
        "dense_rows": dense_rows,
        "avg_non_empty_per_row": round(avg_non_empty_per_row, 2),
        "structure_type": structure_type,
        "top_keyword_hits": dict(keyword_hits_counter.most_common(20)),
        "candidate_rows": candidate_rows[:top_rows_limit],
    }


def _try_pandas_preview(
    *,
    local_path: Path,
    sheet_name: str,
    preview_rows: int,
) -> dict[str, Any]:
    try:
        import pandas as pd

        df_preview = pd.read_excel(
            local_path,
            sheet_name=sheet_name,
            nrows=preview_rows,
        )

        unnamed_count = sum(1 for c in df_preview.columns if str(c).startswith("Unnamed"))
        total_cols = int(df_preview.shape[1])
        unnamed_ratio = round(unnamed_count / total_cols, 3) if total_cols else 0

        return {
            "ok": True,
            "rows_previewed": int(len(df_preview)),
            "columns": [str(c) for c in df_preview.columns.tolist()],
            "shape_preview": [int(df_preview.shape[0]), int(df_preview.shape[1])],
            "unnamed_columns": unnamed_count,
            "unnamed_ratio": unnamed_ratio,
            "preview": df_preview.fillna("").astype(str).to_dict(orient="records"),
            "error": "",
        }

    except Exception as e:
        return {
            "ok": False,
            "rows_previewed": 0,
            "columns": [],
            "shape_preview": [],
            "unnamed_columns": 0,
            "unnamed_ratio": 0,
            "preview": [],
            "error": repr(e),
        }


def _build_excel_analytic_summary(
    *,
    sheet_infos: list[dict[str, Any]],
    question: str,
) -> dict[str, Any]:
    ok_sheets = [s for s in sheet_infos if s.get("ok")]
    ranked = sorted(
        ok_sheets,
        key=lambda s: (
            int(s.get("sheet_relevance_score") or 0),
            int(s.get("cell_scan", {}).get("numeric_cells") or 0),
            int(s.get("cell_scan", {}).get("non_empty_cells") or 0),
        ),
        reverse=True,
    )

    all_hits: Counter[str] = Counter()
    structure_counter: Counter[str] = Counter()
    numeric_total = 0
    monetary_total = 0
    percent_total = 0

    for sheet in ok_sheets:
        scan = sheet.get("cell_scan") or {}
        all_hits.update(scan.get("top_keyword_hits") or {})
        structure_counter[scan.get("structure_type") or "indefinido"] += 1
        numeric_total += int(scan.get("numeric_cells") or 0)
        monetary_total += int(scan.get("monetary_cells") or 0)
        percent_total += int(scan.get("percent_cells") or 0)

    recommended_sheets = [
        {
            "sheet_name": s.get("sheet_name"),
            "score": s.get("sheet_relevance_score"),
            "hits": s.get("sheet_relevance_hits"),
            "structure_type": (s.get("cell_scan") or {}).get("structure_type"),
            "numeric_cells": (s.get("cell_scan") or {}).get("numeric_cells"),
            "candidate_rows_count": len((s.get("cell_scan") or {}).get("candidate_rows") or []),
        }
        for s in ranked[:8]
    ]

    return {
        "question": question,
        "ok_sheets_count": len(ok_sheets),
        "recommended_sheets": recommended_sheets,
        "top_terms_found": dict(all_hits.most_common(20)),
        "structure_counter": dict(structure_counter),
        "numeric_cells_total_scanned": numeric_total,
        "monetary_cells_total_scanned": monetary_total,
        "percent_cells_total_scanned": percent_total,
        "interpretation": _interpret_excel_structure(
            recommended_sheets=recommended_sheets,
            structure_counter=structure_counter,
            numeric_total=numeric_total,
            monetary_total=monetary_total,
            percent_total=percent_total,
        ),
    }


def _interpret_excel_structure(
    *,
    recommended_sheets: list[dict[str, Any]],
    structure_counter: Counter[str],
    numeric_total: int,
    monetary_total: int,
    percent_total: int,
) -> list[str]:
    notes: list[str] = []

    if not recommended_sheets:
        notes.append("Não foi possível identificar abas claramente relevantes para a pergunta.")
        return notes

    top_names = ", ".join(str(s["sheet_name"]) for s in recommended_sheets[:3] if s.get("sheet_name"))
    if top_names:
        notes.append(f"As abas mais promissoras para análise são: {top_names}.")

    if structure_counter:
        dominant = structure_counter.most_common(1)[0][0]
        if dominant == "dashboard_esparso":
            notes.append(
                "A estrutura predominante parece ser de dashboard/relatório formatado, "
                "não de tabela limpa. A leitura por células/blocos é mais adequada que pandas tabular puro."
            )
        elif dominant == "tabela_densa":
            notes.append(
                "A estrutura predominante parece ter blocos tabulares densos. "
                "Há boa chance de extração automática por DataFrame após identificação correta do cabeçalho."
            )
        else:
            notes.append(
                "A estrutura parece mista, com blocos de texto, números e possíveis tabelas."
            )

    if numeric_total > 0:
        notes.append(f"Foram encontrados {numeric_total} valores numéricos na área varrida das abas inspecionadas.")

    if monetary_total > 0:
        notes.append(f"Foram encontrados {monetary_total} valores com aparência monetária.")

    if percent_total > 0:
        notes.append(f"Foram encontrados {percent_total} valores percentuais.")

    notes.append(
        "Esta etapa ainda não consolida indicador final; ela identifica fontes, abas e blocos candidatos para cálculo seguro."
    )

    return notes


def _inspect_excel_file(
    *,
    local_path: Path,
    source: dict[str, Any],
    question: str,
    max_sheets: int = 12,
    preview_rows: int = 5,
    max_scan_rows: int = 140,
    max_scan_cols: int = 45,
) -> FileInspection:
    inspection = FileInspection(
        ok=False,
        file_name=_safe_get(source, "name"),
        file_id=_safe_get(source, "id"),
        path=_safe_get(source, "path"),
        mime_type=_safe_get(source, "mimeType"),
        local_path=str(local_path),
        kind="excel",
    )

    try:
        import pandas as pd
        from openpyxl import load_workbook
    except Exception as e:
        inspection.errors.append(
            f"Dependência ausente para abrir Excel. Instale pandas/openpyxl. Erro: {repr(e)}"
        )
        inspection.message = "Falha: dependência ausente para abrir Excel."
        return inspection

    try:
        excel = pd.ExcelFile(local_path)
        sheet_names = list(excel.sheet_names)

        ranked_sheet_names: list[dict[str, Any]] = []
        for sheet_name in sheet_names:
            score, hits = _rank_sheet_name(sheet_name, question)
            ranked_sheet_names.append(
                {
                    "sheet_name": sheet_name,
                    "score": score,
                    "hits": hits,
                }
            )

        ranked_sheet_names.sort(key=lambda x: int(x["score"]), reverse=True)

        first_sheets = sheet_names[: min(5, len(sheet_names))]
        ranked_sheets = [x["sheet_name"] for x in ranked_sheet_names]
        selected_sheet_names: list[str] = []

        for s in first_sheets + ranked_sheets:
            if s not in selected_sheet_names:
                selected_sheet_names.append(s)
            if len(selected_sheet_names) >= max_sheets:
                break

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(local_path, read_only=True, data_only=True)

        inspection.metadata = {
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "inspected_sheets": selected_sheet_names,
            "ranking_by_sheet_name": ranked_sheet_names[:20],
            "analysis_mode": "cell_blocks_plus_pandas_preview",
        }

        for sheet_name in selected_sheet_names:
            sheet_info: dict[str, Any] = {
                "sheet_name": sheet_name,
                "ok": False,
                "rows_previewed": 0,
                "columns": [],
                "shape_preview": [],
                "unnamed_columns": 0,
                "unnamed_ratio": 0,
                "preview": [],
                "error": "",
                "sheet_relevance_score": 0,
                "sheet_relevance_hits": [],
                "cell_scan": {},
            }

            try:
                relevance_score, relevance_hits = _rank_sheet_name(sheet_name, question)
                sheet_info["sheet_relevance_score"] = relevance_score
                sheet_info["sheet_relevance_hits"] = relevance_hits

                preview = _try_pandas_preview(
                    local_path=local_path,
                    sheet_name=sheet_name,
                    preview_rows=preview_rows,
                )

                sheet_info.update(preview)

                ws = wb[sheet_name]
                cell_scan = _scan_excel_sheet_cells(
                    ws=ws,
                    sheet_name=sheet_name,
                    question=question,
                    max_scan_rows=max_scan_rows,
                    max_scan_cols=max_scan_cols,
                )

                content_bonus = 0
                if cell_scan.get("candidate_rows"):
                    content_bonus += min(len(cell_scan["candidate_rows"]) * 2, 20)
                if cell_scan.get("numeric_cells", 0) > 0:
                    content_bonus += 5
                if cell_scan.get("top_keyword_hits"):
                    content_bonus += min(len(cell_scan["top_keyword_hits"]) * 2, 20)

                sheet_info["sheet_relevance_score"] = int(sheet_info["sheet_relevance_score"]) + content_bonus
                sheet_info["cell_scan"] = cell_scan
                sheet_info["ok"] = True

            except Exception as e:
                sheet_info["error"] = repr(e)

            inspection.sheets.append(sheet_info)

        try:
            wb.close()
        except Exception:
            pass

        inspection.analytic_summary = _build_excel_analytic_summary(
            sheet_infos=inspection.sheets,
            question=question,
        )

        inspection.ok = True
        inspection.message = (
            f"Excel aberto com sucesso. "
            f"{len(sheet_names)} aba(s) encontrada(s); "
            f"{len(inspection.sheets)} aba(s) analisada(s) por prévia tabular e varredura de células."
        )
        return inspection

    except Exception as e:
        inspection.errors.append(repr(e))
        inspection.message = (
            "Falha ao abrir/analisar Excel. Pode ser arquivo muito grande, protegido, "
            "formato legado sem engine instalada, ou estrutura incompatível."
        )
        return inspection


def _inspect_csv_file(
    *,
    local_path: Path,
    source: dict[str, Any],
    question: str,
    preview_rows: int = 20,
) -> FileInspection:
    inspection = FileInspection(
        ok=False,
        file_name=_safe_get(source, "name"),
        file_id=_safe_get(source, "id"),
        path=_safe_get(source, "path"),
        mime_type=_safe_get(source, "mimeType"),
        local_path=str(local_path),
        kind="csv",
    )

    try:
        import pandas as pd
    except Exception as e:
        inspection.errors.append(
            f"pandas não está disponível para abrir CSV. Erro: {repr(e)}"
        )
        inspection.message = "Falha: pandas não disponível."
        return inspection

    try:
        df_preview = pd.read_csv(local_path, nrows=preview_rows, sep=None, engine="python")
        columns = [str(c) for c in df_preview.columns.tolist()]
        text_blob = " ".join(columns + df_preview.fillna("").astype(str).head(preview_rows).to_string().split())
        score, hits = _score_text_for_question(text_blob, question)

        inspection.ok = True
        inspection.metadata = {
            "rows_previewed": int(len(df_preview)),
            "columns": columns,
            "shape_preview": [int(df_preview.shape[0]), int(df_preview.shape[1])],
        }
        inspection.sheets = [
            {
                "sheet_name": "CSV",
                "ok": True,
                "rows_previewed": int(len(df_preview)),
                "columns": columns,
                "shape_preview": [int(df_preview.shape[0]), int(df_preview.shape[1])],
                "preview": df_preview.fillna("").astype(str).to_dict(orient="records"),
                "error": "",
                "sheet_relevance_score": score,
                "sheet_relevance_hits": hits,
                "cell_scan": {},
            }
        ]
        inspection.analytic_summary = {
            "recommended_sheets": [
                {
                    "sheet_name": "CSV",
                    "score": score,
                    "hits": hits,
                    "structure_type": "tabela_densa",
                    "numeric_cells": 0,
                    "candidate_rows_count": 0,
                }
            ],
            "interpretation": [
                "Arquivo CSV aberto com sucesso. A próxima etapa pode aplicar leitura tabular completa."
            ],
        }
        inspection.message = "CSV aberto com sucesso."
        return inspection

    except Exception as e:
        inspection.errors.append(repr(e))
        inspection.message = "Falha ao abrir CSV."
        return inspection


def _inspect_pdf_file(
    *,
    local_path: Path,
    source: dict[str, Any],
    question: str,
    max_pages: int = 5,
    max_chars: int = 5000,
) -> FileInspection:
    inspection = FileInspection(
        ok=False,
        file_name=_safe_get(source, "name"),
        file_id=_safe_get(source, "id"),
        path=_safe_get(source, "path"),
        mime_type=_safe_get(source, "mimeType"),
        local_path=str(local_path),
        kind="pdf",
    )

    text_parts: list[str] = []
    errors: list[str] = []

    try:
        from pypdf import PdfReader

        reader = PdfReader(str(local_path))
        total_pages = len(reader.pages)

        for page_idx in range(min(total_pages, max_pages)):
            page_text = reader.pages[page_idx].extract_text() or ""
            if page_text.strip():
                text_parts.append(page_text.strip())

        inspection.metadata = {
            "pages_total": total_pages,
            "pages_inspected": min(total_pages, max_pages),
            "reader": "pypdf",
        }

    except Exception as e:
        errors.append(f"pypdf falhou: {repr(e)}")

    if not text_parts:
        try:
            import pdfplumber

            with pdfplumber.open(str(local_path)) as pdf:
                total_pages = len(pdf.pages)
                for page in pdf.pages[:max_pages]:
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        text_parts.append(page_text.strip())

            inspection.metadata = {
                "pages_total": total_pages,
                "pages_inspected": min(total_pages, max_pages),
                "reader": "pdfplumber",
            }

        except Exception as e:
            errors.append(f"pdfplumber falhou: {repr(e)}")

    preview = "\n\n".join(text_parts).strip()

    if preview:
        score, hits = _score_text_for_question(preview, question)
        inspection.ok = True
        inspection.text_preview = preview[:max_chars]
        inspection.analytic_summary = {
            "text_relevance_score": score,
            "hits": hits,
            "interpretation": [
                "PDF aberto com sucesso e texto inicial extraído.",
                "A próxima etapa pode aplicar busca por trechos e resumo com citações internas.",
            ],
        }
        inspection.message = "PDF aberto com sucesso e texto inicial extraído."
        inspection.errors = errors
        return inspection

    inspection.errors = errors or ["Não foi possível extrair texto do PDF."]
    inspection.message = "PDF baixado, mas sem texto extraível nesta prévia."
    return inspection


def _inspect_downloaded_file(
    *,
    local_path: Path,
    source: dict[str, Any],
    question: str,
    max_sheets: int,
    preview_rows: int,
    pdf_pages: int,
    max_scan_rows: int,
    max_scan_cols: int,
) -> FileInspection:
    mime = _safe_get(source, "mimeType")
    extension = _normalize_text(source.get("extension", ""))

    if mime in EXCEL_MIMES or extension in {"xlsx", "xlsm", "xls"}:
        return _inspect_excel_file(
            local_path=local_path,
            source=source,
            question=question,
            max_sheets=max_sheets,
            preview_rows=preview_rows,
            max_scan_rows=max_scan_rows,
            max_scan_cols=max_scan_cols,
        )

    if mime == CSV_MIME or extension == "csv":
        return _inspect_csv_file(
            local_path=local_path,
            source=source,
            question=question,
            preview_rows=max(preview_rows, 20),
        )

    if mime == PDF_MIME or extension == "pdf":
        return _inspect_pdf_file(
            local_path=local_path,
            source=source,
            question=question,
            max_pages=pdf_pages,
        )

    return FileInspection(
        ok=False,
        file_name=_safe_get(source, "name"),
        file_id=_safe_get(source, "id"),
        path=_safe_get(source, "path"),
        mime_type=mime,
        local_path=str(local_path),
        kind="unsupported",
        message="Tipo de arquivo ainda não suportado para inspeção automática.",
        errors=[f"MIME não suportado: {mime}"],
    )


def _download_and_inspect_source(
    *,
    source: dict[str, Any],
    question: str,
    download_file,
    cache_dir: str | Path | None,
    max_sheets: int,
    preview_rows: int,
    pdf_pages: int,
    max_scan_rows: int,
    max_scan_cols: int,
) -> FileInspection:
    file_id = _safe_get(source, "id")
    file_name = _safe_get(source, "name", "arquivo_drive")

    if not file_id:
        return FileInspection(
            ok=False,
            file_name=file_name,
            file_id="",
            path=_safe_get(source, "path"),
            mime_type=_safe_get(source, "mimeType"),
            message="Arquivo sem ID no catálogo.",
            errors=["Arquivo sem ID no catálogo."],
        )

    try:
        file_bytes = download_file(file_id)
    except Exception as e:
        return FileInspection(
            ok=False,
            file_name=file_name,
            file_id=file_id,
            path=_safe_get(source, "path"),
            mime_type=_safe_get(source, "mimeType"),
            message="Falha ao baixar arquivo do Google Drive.",
            errors=[repr(e)],
        )

    try:
        local_path = _write_bytes_to_cache(
            file_bytes=file_bytes,
            file_name=file_name,
            file_id=file_id,
            cache_dir=cache_dir,
        )
    except Exception as e:
        return FileInspection(
            ok=False,
            file_name=file_name,
            file_id=file_id,
            path=_safe_get(source, "path"),
            mime_type=_safe_get(source, "mimeType"),
            message="Falha ao salvar arquivo em cache local.",
            errors=[repr(e)],
        )

    return _inspect_downloaded_file(
        local_path=local_path,
        source=source,
        question=question,
        max_sheets=max_sheets,
        preview_rows=preview_rows,
        pdf_pages=pdf_pages,
        max_scan_rows=max_scan_rows,
        max_scan_cols=max_scan_cols,
    )


def _build_safe_response_text(
    *,
    question: str,
    selected_sources: list[dict[str, Any]],
    inspections: list[FileInspection],
    requires_drive: bool,
) -> str:
    lines: list[str] = []

    lines.append("Análise institucional preliminar com base no Google Drive do MAIC.")
    lines.append("")
    lines.append(f"Pergunta analisada: {question}")
    lines.append("")

    if requires_drive:
        lines.append("Regra aplicada: pergunta classificada como institucional/analítica. Usei apenas fontes candidatas do catálogo do Drive.")
    else:
        lines.append("Regra aplicada: pergunta não parece obrigar Drive, mas foi analisada contra o catálogo por solicitação do fluxo.")

    lines.append("")
    lines.append("Fontes selecionadas:")

    for idx, source in enumerate(selected_sources, start=1):
        lines.append(
            f"{idx}. {source.get('name', '')} "
            f"| categoria={source.get('category_detected', '')} "
            f"| área={source.get('area_detected', '')} "
            f"| ano={source.get('year_detected', '')} "
            f"| score={source.get('score', '')} "
            f"| score_ajustado={source.get('adjusted_score', '')}"
        )
        lines.append(f"   Caminho: {source.get('path', '')}")
        lines.append(f"   ID: {source.get('id', '')}")

        if source.get("territorial_notes"):
            lines.append(f"   Ajuste territorial: {', '.join(source.get('territorial_notes', []))}")

    lines.append("")
    lines.append("Leitura analítica dos arquivos:")

    if not inspections:
        lines.append("- Nenhum arquivo foi baixado/inspecionado.")
        return "\n".join(lines)

    for idx, inspection in enumerate(inspections, start=1):
        status = "OK" if inspection.ok else "ERRO"
        lines.append("")
        lines.append(f"{idx}. [{status}] {inspection.file_name}")
        lines.append(f"   Tipo: {inspection.kind or inspection.mime_type}")
        lines.append(f"   Mensagem: {inspection.message}")
        lines.append(f"   Cache local: {inspection.local_path}")

        if inspection.kind in {"excel", "csv"}:
            _append_excel_or_csv_analysis(lines, inspection)

        elif inspection.kind == "pdf":
            _append_pdf_analysis(lines, inspection)

        if inspection.errors:
            lines.append("   Erros/avisos:")
            for err in inspection.errors[:5]:
                lines.append(f"   - {err}")

    lines.append("")
    lines.append("Conclusão operacional:")
    lines.append("- O Drive foi consultado e a análise usou fonte institucional localizada no catálogo.")
    lines.append("- Esta versão já faz varredura por células/blocos, útil para dashboards IMD/RGQC com células mescladas.")
    lines.append("- O ranqueamento territorial foi reforçado para evitar falso positivo como Região Norte versus Asa Norte.")
    lines.append("- A próxima etapa é transformar os blocos candidatos em cálculo final de indicador/custo por regra específica de cada base.")

    return "\n".join(lines)


def _append_excel_or_csv_analysis(lines: list[str], inspection: FileInspection) -> None:
    sheet_count = inspection.metadata.get("sheet_count")
    if sheet_count is not None:
        lines.append(f"   Abas encontradas: {sheet_count}")

    summary = inspection.analytic_summary or {}
    interpretation = summary.get("interpretation") or []

    if interpretation:
        lines.append("   Achados estruturais:")
        for note in interpretation[:6]:
            lines.append(f"   - {note}")

    recommended = summary.get("recommended_sheets") or []
    if recommended:
        lines.append("   Abas/blocos mais promissores:")
        for sheet in recommended[:6]:
            lines.append(
                f"   - {sheet.get('sheet_name')} "
                f"| score={sheet.get('score')} "
                f"| estrutura={sheet.get('structure_type')} "
                f"| células numéricas={sheet.get('numeric_cells')} "
                f"| linhas candidatas={sheet.get('candidate_rows_count')}"
            )

    top_terms = summary.get("top_terms_found") or {}
    if top_terms:
        top_terms_text = ", ".join([f"{k}({v})" for k, v in list(top_terms.items())[:12]])
        lines.append(f"   Termos analíticos encontrados: {top_terms_text}")

    for sheet in inspection.sheets[:5]:
        lines.append(
            f"   - Aba: {sheet.get('sheet_name')} "
            f"| score={sheet.get('sheet_relevance_score')} "
            f"| colunas={len(sheet.get('columns') or [])} "
            f"| linhas prévia={sheet.get('rows_previewed')}"
        )

        columns = sheet.get("columns") or []
        if columns:
            cols_preview = ", ".join(columns[:12])
            if len(columns) > 12:
                cols_preview += f", ... +{len(columns) - 12}"
            lines.append(f"     Colunas: {cols_preview}")

        scan = sheet.get("cell_scan") or {}
        candidate_rows = scan.get("candidate_rows") or []

        if candidate_rows:
            lines.append("     Linhas/blocos candidatos:")
            for row in candidate_rows[:3]:
                values = row.get("values") or []
                values_text = " | ".join(str(v) for v in values[:10])
                lines.append(
                    f"     - linha {row.get('row_index')} "
                    f"| score={row.get('score')} "
                    f"| hits={', '.join(row.get('hits') or [])} "
                    f"| {values_text[:350]}"
                )


def _append_pdf_analysis(lines: list[str], inspection: FileInspection) -> None:
    summary = inspection.analytic_summary or {}

    if summary.get("hits"):
        lines.append(f"   Termos encontrados: {', '.join(summary.get('hits')[:12])}")

    interpretation = summary.get("interpretation") or []
    if interpretation:
        lines.append("   Achados:")
        for note in interpretation:
            lines.append(f"   - {note}")

    if inspection.text_preview:
        preview = inspection.text_preview.replace("\n", " ")
        preview = re.sub(r"\s+", " ", preview).strip()
        lines.append(f"   Prévia textual: {preview[:900]}")


def analyze_question_from_drive(
    question: str,
    *,
    catalog_limit: int = 12,
    selected_limit: int = 1,
    cache_dir: str | Path | None = None,
    max_sheets: int = 12,
    preview_rows: int = 5,
    pdf_pages: int = 5,
    max_scan_rows: int = 140,
    max_scan_cols: int = 45,
    require_institutional_question: bool = False,
) -> dict[str, Any]:
    _bootstrap_import_path()

    from services.drive_catalog import find_best_sources_for_question
    from services.gdrive import download_file

    question = (question or "").strip()

    if not question:
        result = DriveAnalysisResult(
            ok=False,
            requires_drive=False,
            catalog_available=False,
            found_sources=False,
            message="Pergunta vazia.",
            question=question,
            text="Pergunta vazia. Informe uma pergunta sobre monitoramento, indicadores ou custos.",
            errors=["Pergunta vazia."],
        )
        return result.to_dict()

    catalog_response = find_best_sources_for_question(
        question,
        limit=catalog_limit,
    )

    requires_drive = bool(catalog_response.get("requires_drive"))
    catalog_available = bool(catalog_response.get("catalog_available"))
    found = bool(catalog_response.get("found"))
    raw_sources = catalog_response.get("results", []) or []

    if require_institutional_question and not requires_drive:
        result = DriveAnalysisResult(
            ok=False,
            requires_drive=False,
            catalog_available=catalog_available,
            found_sources=found,
            message="Pergunta não classificada como institucional/analítica.",
            question=question,
            selected_sources=[],
            inspections=[],
            text=(
                "Essa pergunta não foi classificada como institucional/analítica. "
                "Ela pode seguir para resposta geral do ChatGPT."
            ),
            errors=[],
        )
        return result.to_dict()

    if not catalog_available:
        msg = catalog_response.get(
            "message",
            "Catálogo do Drive não encontrado ou vazio.",
        )
        result = DriveAnalysisResult(
            ok=False,
            requires_drive=requires_drive,
            catalog_available=False,
            found_sources=False,
            message=msg,
            question=question,
            selected_sources=[],
            inspections=[],
            text=(
                "Não encontrei o catálogo local do Google Drive. "
                "Rode primeiro: python scripts/test_gdrive.py"
            ),
            errors=[msg],
        )
        return result.to_dict()

    if not found or not raw_sources:
        result = DriveAnalysisResult(
            ok=False,
            requires_drive=requires_drive,
            catalog_available=True,
            found_sources=False,
            message="Nenhuma fonte candidata encontrada no catálogo.",
            question=question,
            selected_sources=[],
            inspections=[],
            text=(
                "Não encontrei fonte institucional no Drive para responder com segurança. "
                "Atualize o catálogo com python scripts/test_gdrive.py ou revise a pergunta."
            ),
            errors=[],
        )
        return result.to_dict()

    selected_sources = _rerank_sources_for_question(
        raw_sources,
        question,
        limit=selected_limit,
    )

    inspections: list[FileInspection] = []

    for source in selected_sources:
        inspection = _download_and_inspect_source(
            source=source,
            question=question,
            download_file=download_file,
            cache_dir=cache_dir,
            max_sheets=max_sheets,
            preview_rows=preview_rows,
            pdf_pages=pdf_pages,
            max_scan_rows=max_scan_rows,
            max_scan_cols=max_scan_cols,
        )
        inspections.append(inspection)

    ok = any(i.ok for i in inspections)

    text = _build_safe_response_text(
        question=question,
        selected_sources=selected_sources,
        inspections=inspections,
        requires_drive=requires_drive,
    )

    result = DriveAnalysisResult(
        ok=ok,
        requires_drive=requires_drive,
        catalog_available=True,
        found_sources=True,
        message="Análise institucional preliminar concluída." if ok else "Fontes encontradas, mas houve falha na inspeção.",
        question=question,
        selected_sources=selected_sources,
        inspections=[_inspection_to_dict(i) for i in inspections],
        text=text,
        errors=[err for i in inspections for err in i.errors],
    )

    return result.to_dict()


def analyze_question_text(question: str, **kwargs: Any) -> str:
    result = analyze_question_from_drive(question, **kwargs)
    return str(result.get("text", ""))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Análise institucional usando catálogo + arquivos do Google Drive do MAIC."
    )

    parser.add_argument(
        "question",
        nargs="*",
        help="Pergunta institucional. Exemplo: custos IMD 2024 região oeste",
    )

    parser.add_argument(
        "--catalog-limit",
        type=int,
        default=12,
        help="Quantidade de fontes candidatas vindas do catálogo. Padrão: 12.",
    )

    parser.add_argument(
        "--selected-limit",
        type=int,
        default=1,
        help="Quantidade de arquivos para baixar/analisar. Padrão: 1.",
    )

    parser.add_argument(
        "--max-sheets",
        type=int,
        default=12,
        help="Máximo de abas para analisar em cada Excel. Padrão: 12.",
    )

    parser.add_argument(
        "--preview-rows",
        type=int,
        default=5,
        help="Linhas de prévia tabular por aba. Padrão: 5.",
    )

    parser.add_argument(
        "--pdf-pages",
        type=int,
        default=5,
        help="Páginas iniciais para prévia de PDF. Padrão: 5.",
    )

    parser.add_argument(
        "--max-scan-rows",
        type=int,
        default=140,
        help="Máximo de linhas varridas por aba na leitura por células. Padrão: 140.",
    )

    parser.add_argument(
        "--max-scan-cols",
        type=int,
        default=45,
        help="Máximo de colunas varridas por aba na leitura por células. Padrão: 45.",
    )

    parser.add_argument(
        "--json",
        action="store_true",
        help="Imprime saída completa em JSON.",
    )

    parser.add_argument(
        "--cache-dir",
        default=str(DEFAULT_CACHE_DIR),
        help="Pasta de cache dos arquivos baixados. Padrão: scripts/drive_cache.",
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    question = " ".join(args.question).strip() or "custos IMD 2024 região oeste"

    print("\n=== TESTE DRIVE ANALYTICS / MAIC ===\n")
    print(f"Pergunta: {question}\n")

    result = analyze_question_from_drive(
        question,
        catalog_limit=args.catalog_limit,
        selected_limit=args.selected_limit,
        cache_dir=args.cache_dir,
        max_sheets=args.max_sheets,
        preview_rows=args.preview_rows,
        pdf_pages=args.pdf_pages,
        max_scan_rows=args.max_scan_rows,
        max_scan_cols=args.max_scan_cols,
    )

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(result.get("text", ""))

    print("\n=== FIM ===\n")


if __name__ == "__main__":
    main()